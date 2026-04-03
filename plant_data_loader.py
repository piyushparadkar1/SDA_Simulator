"""
plant_data_loader.py
====================
Loads HPCL PDA plant data from two Excel workbooks and produces
a normalized calibration dataset.

Primary output: build_calibration_dataset() -> dict with two DataFrames:
  'dcs_hourly'    -> all 15,841 hourly DCS rows (for thermal + yield calibration)
  'visc_anchored' -> ~2,000 rows where DCS conditions align with LIMS DAO viscosity
                     (for viscosity calibration)
"""

import logging
import pandas as pd
import numpy as np
from datetime import datetime

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# DCS tag-name to semantic-name mapping
# ---------------------------------------------------------------------------
TAG_MAP = {
    'fc4101a.PV - Average':     'feed_flow_a',
    'fc4101b.PV - Average':     'feed_flow_b',
    '41ti41132.PV - Average':   'feed_temp_a',
    '41ti41133.PV - Average':   'feed_temp_b',
    '41ti41101.PV - Average':   'propane_temp',
    'tc4102a.pv - Average':     't_top_a',
    'ti4107a.pv - Average':     't_steam_coil_a',
    'ti4108a.pv - Average':     't_mid_a',
    'ti4110a.pv - Average':     't_bot_a',
    'tc4102b.pv - Average':     't_top_b',
    'ti4107b.pv - Average':     't_steam_coil_b',
    'ti4108b.pv - Average':     't_mid_b',
    'ti4110b.pv - Average':     't_bot_b',
    '41fic4103a.pv - Average':  'prop_primary_a',
    '41fic4103b.pv - Average':  'prop_primary_b',
    '41fic41113.pv - Average':  'prop_secondary_a',
    '41fic41114.pv - Average':  'prop_secondary_b',
    '41fic41110.pv - Average':  'dao_flow',
    '41fi4119.pv - Average':    'asphalt_flow',
}


def _clean_dcs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove bad rows and fix instrument faults in DCS data.

    Filters (in order):
    1. Both extractors offline (feed_flow_a < 2 AND feed_flow_b < 2) → drop row.
    2. Temperature instrument faults (> 150°C or < 10°C while not zero) → set to NaN.
    3. Train offline while feed flowing (top_temp < 30°C at flow > 2) → that train's temps to NaN.
    4. Negative flows → clamp to 0.
    5. Both product flows zero while feed flowing → drop row (measurement fault, not shutdown).

    Parameters
    ----------
    df : DataFrame with DatetimeIndex and semantic column names from TAG_MAP.

    Returns
    -------
    Cleaned DataFrame (some rows removed, some values set to NaN).
    """
    n_before = len(df)

    # 1. Both extractors offline
    fa = df['feed_flow_a'].fillna(0) if 'feed_flow_a' in df.columns else pd.Series(0.0, index=df.index)
    fb = df['feed_flow_b'].fillna(0) if 'feed_flow_b' in df.columns else pd.Series(0.0, index=df.index)
    mask_offline = (fa < 2) & (fb < 2)
    df = df[~mask_offline].copy()

    # 2. Temperature instrument faults → NaN (keep row; other train may be OK)
    temp_cols = [c for c in df.columns
                 if c.startswith(('t_top_', 't_mid_', 't_bot_', 't_steam_coil_'))]
    for col in temp_cols:
        df.loc[df[col] > 150, col] = np.nan
        df.loc[(df[col] < 10) & df[col].notna() & (df[col] != 0), col] = np.nan

    # 3. Train offline while feed flowing → that train's temps to NaN
    for train in ['a', 'b']:
        feed_col = f'feed_flow_{train}'
        top_col  = f't_top_{train}'
        if feed_col in df.columns and top_col in df.columns:
            mask_train_off = (df[feed_col] > 2) & (df[top_col].fillna(0) < 30)
            t_train_cols = [c for c in df.columns
                            if c.endswith(f'_{train}') and
                            c.startswith(('t_top_', 't_mid_', 't_bot_', 't_steam_coil_'))]
            for col in t_train_cols:
                df.loc[mask_train_off, col] = np.nan

    # 4. Negative flows → 0
    flow_cols = [c for c in df.columns
                 if 'flow' in c.lower() or c.startswith('prop_')]
    for col in flow_cols:
        df.loc[df[col] < 0, col] = 0

    # 5. Both product flows zero while feed flowing → bad data, drop
    if 'dao_flow' in df.columns and 'asphalt_flow' in df.columns:
        fa2 = df['feed_flow_a'].fillna(0) if 'feed_flow_a' in df.columns else pd.Series(0.0, index=df.index)
        fb2 = df['feed_flow_b'].fillna(0) if 'feed_flow_b' in df.columns else pd.Series(0.0, index=df.index)
        mask_no_product = (
            (df['dao_flow'].fillna(0)     <= 0) &
            (df['asphalt_flow'].fillna(0) <= 0) &
            ((fa2 > 5) | (fb2 > 5))
        )
        df = df[~mask_no_product]

    n_after = len(df)
    logger.info(f"  DCS cleaning: {n_before} → {n_after} rows "
                f"(removed {n_before - n_after} = {(n_before - n_after) / max(n_before, 1) * 100:.1f}%)")
    return df


def _parse_simple_lims_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Parse a simple LIMS sheet with structure: row 0 = header, rows 1+ = data.

    Finds the date column by searching for 'date' in the header row,
    and uses the last column as the value column.

    Parameters
    ----------
    df : Raw DataFrame read with header=None (so row 0 is the column-name row).

    Returns
    -------
    DataFrame with columns [timestamp, value], sorted ascending, NaN rows dropped.
    """
    header = [str(c).strip().lower() for c in df.iloc[0].values]

    date_idx = next((i for i, h in enumerate(header) if 'date' in h), None)
    value_idx = len(header) - 1   # last column is the result value

    if date_idx is None:
        raise ValueError(f"No date column found in LIMS sheet header: {header}")

    data = df.iloc[1:].copy()
    result = pd.DataFrame({
        'timestamp': pd.to_datetime(data.iloc[:, date_idx], errors='coerce'),
        'value':     pd.to_numeric(data.iloc[:, value_idx], errors='coerce'),
    })
    result = result.dropna(subset=['timestamp', 'value'])
    return result.sort_values('timestamp').reset_index(drop=True)


def _add_steady_state_flag(df: pd.DataFrame,
                           window_hours: int = 2,
                           T_std_max: float = 1.5,
                           flow_std_max: float = 2.0) -> pd.DataFrame:
    """
    Add 'steady_state' boolean column based on rolling std over a trailing window.

    Criteria (per train, ANDed across both trains):
    - Max bed-temperature std < T_std_max (°C) over trailing window_hours
    - Feed flow std < flow_std_max (m³/hr) over trailing window_hours

    Parameters
    ----------
    df : DataFrame with DatetimeIndex (required for time-based rolling window).
    window_hours : Rolling window size in hours.
    T_std_max : Max allowable temperature std [°C].
    flow_std_max : Max allowable feed-flow std [m³/hr].

    Returns
    -------
    df with 'steady_state' boolean column added in-place.
    """
    def _rolling_std(series, window='2h', min_periods=2):
        return series.rolling(window, min_periods=min_periods).std()

    window_str = f'{window_hours}h'
    steady_flags = []

    for train in ['a', 'b']:
        t_cols = [f't_bot_{train}', f't_mid_{train}', f't_top_{train}']
        flow_col = f'feed_flow_{train}'
        t_cols_present = [c for c in t_cols if c in df.columns]

        if not t_cols_present or flow_col not in df.columns:
            steady_flags.append(pd.Series(True, index=df.index))
            continue

        t_stds = pd.concat(
            [_rolling_std(df[c], window=window_str) for c in t_cols_present], axis=1
        ).max(axis=1)
        flow_std = _rolling_std(df[flow_col], window=window_str)

        steady = (t_stds < T_std_max) & (flow_std < flow_std_max)
        steady = steady.fillna(False)
        steady_flags.append(steady)

    if len(steady_flags) == 2:
        df['steady_state'] = steady_flags[0] & steady_flags[1]
    elif steady_flags:
        df['steady_state'] = steady_flags[0]
    else:
        df['steady_state'] = True

    n_steady = int(df['steady_state'].sum())
    logger.info(f"  Steady-state: {n_steady}/{len(df)} rows ({n_steady / max(len(df), 1) * 100:.1f}%)")
    return df


def load_dcs_workbook(filepath: str) -> pd.DataFrame:
    """
    Load extractor_parameters.xlsx -> normalized hourly DataFrame.

    Steps:
    1. Load sheet 'pda_data' with openpyxl engine, header at row index 1 (0-based).
    2. Rename columns using TAG_MAP. Columns not in TAG_MAP are dropped silently.
    3. Drop rows where Timestamp is NaT or feed_flow_a <= 0 or feed_flow_b <= 0.
    4. Compute derived columns: so_ratio_a/b, predilution_frac_a/b,
       feed_flow_total, dao_yield_vol_pct, per-bed T columns.
    5. Compute steady_state flag from rolling std over 2-hr window.
    6. Compute train_valid_a and train_valid_b flags.
    7. Return DataFrame with Timestamp as DatetimeIndex, sorted ascending.
    """
    import pandas as pd

    # --- Load ---
    try:
        import openpyxl as _oxl
        _wb2 = _oxl.load_workbook(filepath, read_only=True)
        _dcs_sheets = _wb2.sheetnames
        _wb2.close()
    except Exception:
        _dcs_sheets = []

    # Fuzzy-match 'pda_data' (strips whitespace + lowercases)
    _pda_sheet = next(
        (s for s in _dcs_sheets if s.strip().lower() == 'pda_data'), None
    )
    if _pda_sheet is None:
        raise ValueError(
            f"Sheet 'pda_data' not found in extractor_parameters.xlsx. "
            f"Available sheets: {_dcs_sheets}"
        )
    try:
        raw = pd.read_excel(filepath, sheet_name=_pda_sheet, header=1, engine='openpyxl')
    except Exception as e:
        raise ValueError(
            f"Cannot load sheet '{_pda_sheet}' from extractor_parameters.xlsx: {e}"
        )

    logger.info(f"DCS raw shape: {raw.shape}")

    # Strip whitespace from column names and build a case-insensitive map
    raw.columns = raw.columns.astype(str).str.strip()

    # Identify and keep the Timestamp column (Column B after dropping empty Column A)
    # After header=1, Column A (empty) becomes 'Unnamed: 0'; Column B becomes whatever
    # the header text is. We identify Timestamp by looking for datetime-like values.
    timestamp_col = None
    for col in raw.columns:
        if col.lower() == 'timestamp':
            timestamp_col = col
            break
    if timestamp_col is None:
        # Try to find by dtype or content
        for col in raw.columns:
            sample = raw[col].dropna().head(10)
            if pd.api.types.is_datetime64_any_dtype(sample):
                timestamp_col = col
                break
    if timestamp_col is None:
        # Fallback: look for the second non-Unnamed column
        non_unnamed = [c for c in raw.columns if not c.startswith('Unnamed')]
        if non_unnamed:
            timestamp_col = non_unnamed[0]
    if timestamp_col is None:
        raise ValueError(
            "Cannot identify Timestamp column in 'pda_data' sheet. "
            "Expected column name 'Timestamp' in row 2."
        )

    # Build reverse map: strip and lower both keys for robust matching
    tag_map_norm = {k.strip().lower(): v for k, v in TAG_MAP.items()}
    rename_map = {}
    for col in raw.columns:
        cl = col.strip().lower()
        if cl in tag_map_norm:
            rename_map[col] = tag_map_norm[cl]

    # Keep only Timestamp + recognized tag columns
    keep_cols = [timestamp_col] + [c for c in raw.columns if c in rename_map]
    df = raw[keep_cols].copy()
    df = df.rename(columns=rename_map)
    df = df.rename(columns={timestamp_col: 'Timestamp'})

    # Coerce Timestamp to datetime
    df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')

    # --- Quality filters ---
    n_raw = len(df)
    df = df.dropna(subset=['Timestamp'])
    logger.info(f"Dropped {n_raw - len(df)} rows with NaT timestamps")

    # Drop rows where BOTH trains offline — single-train operation is preserved;
    # per-train validity is handled by train_valid_a/b flags later.
    n_before = len(df)
    if 'feed_flow_a' in df.columns and 'feed_flow_b' in df.columns:
        both_offline = (df['feed_flow_a'].fillna(0) <= 0) & (df['feed_flow_b'].fillna(0) <= 0)
        df = df[~both_offline]
    logger.info(f"Dropped {n_before - len(df)} rows where both trains offline (feed_flow <= 0)")

    # Sort and set index
    df = df.sort_values('Timestamp').reset_index(drop=True)
    df = df.set_index('Timestamp')
    df.index = pd.DatetimeIndex(df.index)

    # Ensure numeric
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # --- Data quality cleaning (instrument faults, temperature anomalies, offline trains) ---
    df = _clean_dcs(df)

    # --- Derived columns ---
    for train in ['a', 'b']:
        pp = f'prop_primary_{train}'
        ps = f'prop_secondary_{train}'
        ff = f'feed_flow_{train}'

        prop_total = df.get(pp, pd.Series(0.0, index=df.index)).fillna(0) + \
                     df.get(ps, pd.Series(0.0, index=df.index)).fillna(0)
        flow = df.get(ff, pd.Series(1.0, index=df.index)).replace(0, np.nan)

        df[f'so_ratio_{train}'] = prop_total / flow
        df[f'predilution_frac_{train}'] = (
            df.get(ps, pd.Series(0.0, index=df.index)).fillna(0) /
            prop_total.replace(0, np.nan)
        )

    if 'feed_flow_a' in df.columns and 'feed_flow_b' in df.columns:
        df['feed_flow_total'] = df['feed_flow_a'].fillna(0) + df['feed_flow_b'].fillna(0)
        df['feed_flow_total'] = df['feed_flow_total'].replace(0, np.nan)

        if 'dao_flow' in df.columns:
            df['dao_yield_vol_pct'] = df['dao_flow'] / df['feed_flow_total'] * 100.0

    # Filter unrealistic S/O ratios (propane SDA: 2–20 vol/vol)
    for train in ['a', 'b']:
        so_col = f'so_ratio_{train}'
        if so_col in df.columns:
            n_bad = int(((df[so_col] < 2) | (df[so_col] > 20)).sum())
            df.loc[(df[so_col] < 2) | (df[so_col] > 20), so_col] = np.nan
            logger.info(f"  S/O ratio {train}: set {n_bad} out-of-range values to NaN (valid: 2–20)")

    # Filter unrealistic DAO yields
    if 'dao_yield_vol_pct' in df.columns:
        n_bad = int(((df['dao_yield_vol_pct'] < 1) | (df['dao_yield_vol_pct'] > 50)).sum())
        df.loc[(df['dao_yield_vol_pct'] < 1) | (df['dao_yield_vol_pct'] > 50),
               'dao_yield_vol_pct'] = np.nan
        logger.info(f"  DAO yield: set {n_bad} out-of-range values to NaN (valid: 1–50 vol%)")

    # --- Steady-state flag ---
    df = _add_steady_state_flag(df)

    # --- Train validity flags ---
    for train in ['a', 'b']:
        cond = pd.Series(True, index=df.index)

        ff = f'feed_flow_{train}'
        if ff in df.columns:
            cond &= df[ff] > 5.0

        so = f'so_ratio_{train}'
        if so in df.columns:
            cond &= df[so].between(2.0, 18.0)

        tt = f't_top_{train}'
        if tt in df.columns:
            cond &= df[tt].between(60.0, 110.0)

        tb = f't_bot_{train}'
        if tb in df.columns:
            cond &= df[tb].between(50.0, 100.0)

        pdf = f'predilution_frac_{train}'
        if pdf in df.columns:
            cond &= df[pdf].between(0.0, 0.8)

        # Drop rows with NaN in any required column
        cond = cond.fillna(False)
        df[f'train_valid_{train}'] = cond

    n_valid = int((df['train_valid_a'] & df['train_valid_b']).sum())
    logger.info(f"Both-trains-valid rows: {n_valid} / {len(df)}")

    return df


def load_lims_workbook(filepath: str) -> dict:
    """
    Load lims.xlsx -> dict of time-series DataFrames, one per property.

    Returns:
        {
          'feed_visc_135':  DataFrame with columns [timestamp, value]
          'feed_density':   DataFrame with columns [timestamp, value]  (g/cm3)
          'feed_ccr':       DataFrame with columns [timestamp, value]
          'dao_visc_100':   DataFrame with columns [timestamp, value]
          'dao_ccr':        DataFrame with columns [timestamp, value]
          'dao_asphaltene': DataFrame with columns [timestamp, value]
        }
    """

    # --- Get actual sheet names once (handles trailing spaces, mixed case, etc.) ---
    try:
        import openpyxl as _xl
        _wb = _xl.load_workbook(filepath, read_only=True)
        _actual_sheets = _wb.sheetnames
        _wb.close()
    except Exception:
        _actual_sheets = []

    def _resolve_sheet(wanted: str) -> str:
        """Return the real sheet name in the workbook that matches `wanted`.

        Matching is done by stripping and lowercasing both sides.
        Raises ValueError with helpful message if not found.
        """
        wanted_norm = wanted.strip().lower()
        for real in _actual_sheets:
            if real.strip().lower() == wanted_norm:
                return real
        raise ValueError(
            f"Sheet '{wanted}' not found in lims.xlsx. "
            f"Available sheets: {_actual_sheets}"
        )

    def _find_col(df: pd.DataFrame, *keywords) -> str:
        """Find first column whose name contains any of the keywords (case-insensitive)."""
        for kw in keywords:
            kw_lower = kw.lower()
            for col in df.columns:
                if kw_lower in str(col).lower():
                    return col
        raise ValueError(
            f"Cannot find column matching {keywords}. Available: {list(df.columns)}"
        )

    def _load_simple_sheet(filepath, sheet_name, date_kw, value_kw,
                           skiprows=None, header=1) -> pd.DataFrame:
        """Load a simple two-column LIMS sheet (date + value)."""
        real_name = _resolve_sheet(sheet_name)
        try:
            if skiprows is not None:
                raw = pd.read_excel(filepath, sheet_name=real_name,
                                    header=header, skiprows=skiprows, engine='openpyxl')
            else:
                raw = pd.read_excel(filepath, sheet_name=real_name,
                                    header=header, engine='openpyxl')
        except Exception as e:
            raise ValueError(
                f"Cannot load sheet '{real_name}' from lims.xlsx: {e}"
            )
        raw.columns = raw.columns.astype(str).str.strip()

        date_col  = _find_col(raw, *date_kw)
        value_col = _find_col(raw, *value_kw)

        result = pd.DataFrame({
            'timestamp': pd.to_datetime(raw[date_col], errors='coerce'),
            'value':     pd.to_numeric(raw[value_col], errors='coerce'),
        })
        n_before = len(result)
        result = result.dropna(subset=['timestamp', 'value'])
        result = result.sort_values('timestamp').reset_index(drop=True)
        logger.info(
            f"  Sheet '{sheet_name}': loaded {len(result)} rows "
            f"(dropped {n_before - len(result)} invalid)"
        )
        return result

    logger.info("Loading LIMS workbook...")

    # --- feed kin viscosity at 135 degC ---
    feed_visc = _load_simple_sheet(
        filepath, 'feed kin viscosity at 135 degC',
        date_kw=('Sampled Date', 'sampled'),
        value_kw=('135', 'viscosity', 'kinemantoc', 'kinematic'),
    )
    # Filter outliers: 6 zeros and 3 values >1000 cSt confirmed in data audit
    n_v = len(feed_visc)
    feed_visc = feed_visc[feed_visc['value'].between(50, 800)].reset_index(drop=True)
    logger.info(f"  Feed viscosity outlier filter: {n_v} → {len(feed_visc)} rows (kept 50–800 cSt)")

    # --- feed density ---
    feed_density = _load_simple_sheet(
        filepath, 'feed desity at 15degC',
        date_kw=('Sampled Date', 'sampled'),
        value_kw=('Result Value', 'result'),
    )
    # Convert kg/m3 to g/cm3
    if feed_density['value'].mean() > 10:   # still in kg/m3
        feed_density['value'] = feed_density['value'] / 1000.0
        logger.info("  Converted feed_density from kg/m3 to g/cm3")

    # --- feed CCR ---
    feed_ccr = _load_simple_sheet(
        filepath, 'feed ccr wt%',
        date_kw=('Sampled Date', 'sampled'),
        value_kw=('Result Value', 'result'),
    )

    # --- DAO viscosity ---
    dao_visc = _load_simple_sheet(
        filepath, 'dao viscosity',
        date_kw=('Sampled Date', 'sampled'),
        value_kw=('100', 'dao viscosity', 'kinematic'),
    )

    # --- lims pda common (DAO CCR + DAO asphaltene) ---
    dao_ccr = pd.DataFrame({'timestamp': pd.Series(dtype='datetime64[ns]'),
                            'value': pd.Series(dtype='float64')})
    dao_asph = pd.DataFrame({'timestamp': pd.Series(dtype='datetime64[ns]'),
                             'value': pd.Series(dtype='float64')})
    try:
        _common_sheet = _resolve_sheet('lims pda common')
        common_raw = pd.read_excel(
            filepath, sheet_name=_common_sheet, header=1, engine='openpyxl'
        )
        common_raw.columns = common_raw.columns.astype(str).str.strip()

        prod_col  = _find_col(common_raw, 'Product', 'product')
        comp_col  = _find_col(common_raw, 'Component name', 'component')
        res_col   = _find_col(common_raw, 'Result text', 'result text', 'result')
        date_col  = _find_col(common_raw, 'Date result', 'authorised', 'date')

        # Filter to D_DAO only
        dao_mask = common_raw[prod_col].astype(str).str.strip() == 'D_DAO'
        dao_df = common_raw[dao_mask].copy()

        def _extract_common(df, comp_values):
            comp_mask = df[comp_col].astype(str).str.strip().isin(comp_values)
            sub = df[comp_mask].copy()
            values = pd.to_numeric(sub[res_col], errors='coerce')
            timestamps = pd.to_datetime(sub[date_col], errors='coerce')
            result = pd.DataFrame({'timestamp': timestamps.values,
                                   'value': values.values})
            result = result.dropna(subset=['timestamp', 'value'])
            result = result.sort_values('timestamp').reset_index(drop=True)
            return result

        dao_ccr = _extract_common(
            dao_df,
            ['Carbon Residue, Micro method', 'Ramsbottom Carbon Residue',
             'Carbon Residue', 'CCR', 'Conradson Carbon Residue']
        )
        dao_asph = _extract_common(
            dao_df,
            ['Asphaltene', 'Asphaltenes', 'Pentane Insoluble',
             'C5 Insolubles', 'Asphaltene Content']
        )
        logger.info(
            f"  Sheet 'lims pda common': dao_ccr={len(dao_ccr)} rows, "
            f"dao_asphaltene={len(dao_asph)} rows"
        )
    except Exception as e:
        logger.warning(f"Could not load 'lims pda common': {e}")

    return {
        'feed_visc_135':  feed_visc,
        'feed_density':   feed_density,
        'feed_ccr':       feed_ccr,
        'dao_visc_100':   dao_visc,
        'dao_ccr':        dao_ccr,
        'dao_asphaltene': dao_asph,
    }


def _attach_lims_to_dcs(
    dcs: pd.DataFrame,
    lims_series: pd.DataFrame,
    col_name: str,
    tolerance_hours: float,
    lag_hours: float = 0.0,
) -> pd.DataFrame:
    """
    Attach a LIMS time-series to the DCS DataFrame using merge_asof.

    For each DCS row at time T, find the most recent LIMS sample
    at time (T - lag_hours) within tolerance_hours.
    Adds col_name column and col_name + '_age_hr' (staleness indicator)
    and col_name + '_is_stale' (bool: age > tolerance).

    lag_hours: how many hours before T to search for the matching sample.
    Used for DAO properties (sample at T reflects conditions at T - lag).
    """
    if lims_series.empty:
        dcs_out = dcs.copy()
        dcs_out[col_name] = np.nan
        dcs_out[col_name + '_age_hr'] = np.nan
        dcs_out[col_name + '_is_stale'] = True
        return dcs_out

    # Build shifted search timestamps: for each DCS row at T, search at (T - lag)
    search_ts = dcs.index - pd.Timedelta(hours=lag_hours)
    left_df = pd.DataFrame({'_search_ts': search_ts}, index=dcs.index)
    left_df = left_df.reset_index()          # columns: Timestamp, _search_ts
    left_df = left_df.sort_values('_search_ts').reset_index(drop=True)

    right_df = (lims_series
                .sort_values('timestamp')
                .rename(columns={'timestamp': '_lims_ts', 'value': col_name})
                .reset_index(drop=True))

    merged = pd.merge_asof(
        left_df[['Timestamp', '_search_ts']],
        right_df[['_lims_ts', col_name]],
        left_on='_search_ts',
        right_on='_lims_ts',
        direction='backward',
        tolerance=pd.Timedelta(hours=tolerance_hours),
    )

    # Re-align to original DCS index
    merged = merged.set_index('Timestamp')

    dcs_out = dcs.copy()
    dcs_out[col_name] = merged[col_name]

    # Age = search_ts - lims_ts (positive: how old the sample is)
    age_td = (
        pd.Series(search_ts, index=dcs.index) -
        merged['_lims_ts'].reindex(dcs.index)
    )
    dcs_out[col_name + '_age_hr'] = age_td.dt.total_seconds() / 3600.0
    dcs_out[col_name + '_is_stale'] = (
        dcs_out[col_name + '_age_hr'].fillna(float('inf')) > tolerance_hours
    )

    n_matched = int(dcs_out[col_name].notna().sum())
    logger.info(
        f"  Attached '{col_name}': {n_matched} / {len(dcs_out)} DCS rows matched "
        f"(tolerance={tolerance_hours}h, lag={lag_hours}h)"
    )
    return dcs_out


def build_calibration_dataset(
    dcs_filepath: str,
    lims_filepath: str,
) -> dict:
    """
    Main entry point. Loads both workbooks, attaches LIMS to DCS rows,
    applies quality filters, and returns two DataFrames.

    Returns:
        {
          'dcs_hourly':    pd.DataFrame,   # all rows with LIMS columns attached
          'visc_anchored': pd.DataFrame,   # filtered rows for viscosity calibration
          'dataset_info':  dict,           # summary statistics for UI display
        }
    """
    logger.info("=" * 60)
    logger.info("Loading DCS workbook...")
    dcs = load_dcs_workbook(dcs_filepath)
    logger.info(f"DCS loaded: {len(dcs)} rows, "
                f"{dcs.index.min().date()} to {dcs.index.max().date()}")

    logger.info("Loading LIMS workbook...")
    lims = load_lims_workbook(lims_filepath)

    # Attach LIMS to DCS
    logger.info("Attaching LIMS time-series to DCS...")
    dcs = _attach_lims_to_dcs(dcs, lims['feed_visc_135'], 'feed_visc_135',
                               tolerance_hours=12.0, lag_hours=0.0)
    dcs = _attach_lims_to_dcs(dcs, lims['feed_density'],  'feed_density',
                               tolerance_hours=168.0, lag_hours=0.0)
    dcs = _attach_lims_to_dcs(dcs, lims['feed_ccr'],      'feed_ccr',
                               tolerance_hours=168.0, lag_hours=0.0)
    # tolerance=24h: LIMS dao_visc_100 is sampled ~3.5×/day (every ~7h).
    # The old 6h window was too narrow — many samples fell outside and were
    # silently dropped. 24h captures all daily samples (lag_hours=3h already
    # accounts for the physical DCS→LIMS delay).
    dcs = _attach_lims_to_dcs(dcs, lims['dao_visc_100'],  'dao_visc_100',
                               tolerance_hours=24.0, lag_hours=3.0)
    dcs = _attach_lims_to_dcs(dcs, lims['dao_ccr'],       'dao_ccr',
                               tolerance_hours=12.0, lag_hours=3.0)
    dcs = _attach_lims_to_dcs(dcs, lims['dao_asphaltene'],'dao_asphaltene',
                               tolerance_hours=12.0, lag_hours=3.0)

    # Forward-fill sparse LIMS properties with engineering defaults.
    # Feed density (14 measurements) and CCR (10 measurements) must be filled
    # to avoid losing the majority of DCS rows in the calibration dataset.
    # Do NOT forward-fill dao_visc_100 — it is the primary calibration target.
    if 'feed_density' in dcs.columns:
        dcs['feed_density'] = dcs['feed_density'].ffill().fillna(1.028)
        logger.info("  feed_density: forward-filled + default 1.028 g/cm³")
    if 'feed_ccr' in dcs.columns:
        dcs['feed_ccr'] = dcs['feed_ccr'].ffill().fillna(22.8)
        logger.info("  feed_ccr: forward-filled + default 22.8 wt%")
    if 'feed_visc_135' in dcs.columns:
        dcs['feed_visc_135'] = dcs['feed_visc_135'].ffill()
        logger.info("  feed_visc_135: forward-filled (no default — remains NaN at leading edge)")

    # --- Build visc_anchored subset ---
    valid_both = dcs.get('train_valid_a', pd.Series(True, index=dcs.index)) & \
                 dcs.get('train_valid_b', pd.Series(True, index=dcs.index))
    steady     = dcs.get('steady_state', pd.Series(True, index=dcs.index))
    visc_ok    = dcs['dao_visc_100'].notna() & ~dcs['dao_visc_100_is_stale']
    visc_135_ok= dcs['feed_visc_135'].notna()
    density_ok = dcs['feed_density'].notna()

    visc_mask = valid_both & steady & visc_ok & visc_135_ok & density_ok
    visc_anchored = dcs[visc_mask].copy()

    logger.info(
        f"visc_anchored: {len(visc_anchored)} rows "
        f"(valid_both={int(valid_both.sum())}, steady={int(steady.sum())}, "
        f"visc_ok={int(visc_ok.sum())})"
    )
    print(f"[DATA] visc_anchored rows after tolerance=24h fix: {len(visc_anchored)} "
          f"(dao_visc_100 matched: {int(visc_ok.sum())})", flush=True)

    # --- Dataset info ---
    dataset_info = {
        'dcs_rows_total':     int(len(dcs)),
        'dcs_rows_valid':     int(valid_both.sum()),
        'dcs_rows_steady':    int(steady.sum()),
        'visc_anchored_rows': int(len(visc_anchored)),
        'date_range_start':   str(dcs.index.min().date()),
        'date_range_end':     str(dcs.index.max().date()),
        'dao_visc_count':     int(visc_ok.sum()),
        'feed_visc_count':    int(visc_135_ok.sum()),
        'feed_density_count': int(density_ok.sum()),
        'feed_ccr_count':     int(dcs['feed_ccr'].notna().sum()),
    }

    logger.info("Dataset summary:")
    for k, v in dataset_info.items():
        logger.info(f"  {k}: {v}")

    return {
        'dcs_hourly':    dcs,
        'visc_anchored': visc_anchored,
        'dataset_info':  dataset_info,
    }
